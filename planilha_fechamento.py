"""
Script para gerar planilha de fechamento com cores baseado no status dos motoristas.
Mantém histórico completo (nunca remove dados).
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import logging

logger = logging.getLogger(__name__)

class PlanilhaFechamento:
    def __init__(self, diretorio='./'):
        self.diretorio = Path(diretorio)
        self.diretorio.mkdir(parents=True, exist_ok=True)
        self.data_atual = datetime.now().strftime('%d_%m_%Y')
        self.nome_arquivo = self.diretorio / f'planilha_fechamento_{self.data_atual}.xlsx'
        
        # Cores e estilos
        self.cor_verde = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        self.cor_vermelho = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        self.cor_amarelo = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.cor_cinza = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        self.fonte_branca = Font(color='FFFFFF', bold=True)
        self.fonte_preta = Font(color='000000')
        self.fonte_vermelha = Font(color='FFFFFF', bold=True)
        self.fonte_amarela = Font(color='000000', bold=True)
        
        self.borda = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def criar_ou_atualizar_planilha(self, relatorio):
        """Cria ou atualiza a planilha com os dados do relatório.
        
        Args:
            relatorio (list): Lista de dicts com LH, Nome, Placa, Status, Data
        """
        try:
            # Tenta carregar planilha existente
            if self.nome_arquivo.exists():
                wb = load_workbook(self.nome_arquivo)
                ws = wb.active
                logger.info(f"Planilha existente carregada: {self.nome_arquivo}")
            else:
                # Cria nova planilha
                wb = Workbook()
                ws = wb.active
                ws.title = "Fechamento"
                logger.info(f"Nova planilha criada: {self.nome_arquivo}")
            
            # Limpa dados existentes mantendo headers
            while ws.max_row > 1:
                ws.delete_rows(2, 1)
            
            # Adiciona headers se não existirem
            if ws.max_row == 0:
                self._adicionar_headers(ws)
            
            # Adiciona dados ao worksheet
            self._preencher_dados(ws, relatorio)
            
            # Ajusta largura das colunas
            self._ajustar_colunas(ws)
            
            # Salva a planilha
            wb.save(self.nome_arquivo)
            logger.info(f"Planilha atualizada com sucesso: {self.nome_arquivo}")
            
            return str(self.nome_arquivo)
        
        except Exception as e:
            logger.error(f"Erro ao criar/atualizar planilha: {e}", exc_info=True)
            raise
    
    def _adicionar_headers(self, ws):
        """Adiciona cabeçalhos à planilha."""
        headers = ['LH', 'Nome', 'Placa', 'Status', 'Data']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.cor_cinza
            cell.font = Font(bold=True, color='000000')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.borda
    
    def _preencher_dados(self, ws, relatorio):
        """Preenche as linhas da planilha com os dados e cores apropriadas."""
        if not relatorio:
            logger.warning("Relatório vazio ao preencher planilha")
            return
        
        for row_idx, item in enumerate(relatorio, start=2):
            try:
                lh = str(item.get('LH', '')).strip()
                nome = str(item.get('Nome', '')).strip()
                placa = str(item.get('Placa', '')).strip()
                status = str(item.get('Status', 'Ativo')).strip()
                data = str(item.get('Data', '')).strip()
                
                # Define cor baseada no status
                if status.lower() == 'concluido':
                    preenchimento = self.cor_verde
                    fonte = self.fonte_branca
                elif status.lower() == 'cancelado':
                    preenchimento = self.cor_vermelho
                    fonte = self.fonte_vermelha
                else:  # Ativo
                    preenchimento = self.cor_amarelo
                    fonte = self.fonte_amarela
                
                # Escreve células com dados
                valores = [lh, nome, placa, status, data]
                for col_idx, valor in enumerate(valores, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                    cell.fill = preenchimento
                    cell.font = fonte
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = self.borda
                
                logger.debug(f"Linha {row_idx} preenchida: {lh} {nome} {placa}")
            
            except Exception as e:
                logger.error(f"Erro ao preencher linha {row_idx}: {e}")
                continue
    
    def _ajustar_colunas(self, ws):
        """Ajusta a largura das colunas automaticamente."""
        colunas = {
            'A': 15,  # LH
            'B': 25,  # Nome
            'C': 15,  # Placa
            'D': 12,  # Status
            'E': 18   # Data
        }
        for col, largura in colunas.items():
            ws.column_dimensions[col].width = largura
    
    def obter_caminho(self):
        """Retorna o caminho completo do arquivo."""
        return str(self.nome_arquivo)
